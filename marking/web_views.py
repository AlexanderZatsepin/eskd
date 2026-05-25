import json
from collections import defaultdict
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from marking.models import Cabinet, Project, WireEndpoint


@login_required(login_url="/admin/login/")
def cambrics_report(request):
    projects = list(Project.objects.prefetch_related("cabinets").order_by("project_code"))
    context = {
        "projects": projects,
        "cabinets_json": json.dumps(_cabinet_map(projects), ensure_ascii=False),
    }

    if request.method == "POST":
        cabinet_pk = request.POST.get("cabinet_id")
        try:
            cabinet = Cabinet.objects.select_related("project").get(pk=cabinet_pk)
        except Cabinet.DoesNotExist:
            context["error"] = "Шкаф не найден."
        else:
            return _build_cambrics_response(cabinet)

    return render(request, "marking/cambrics_report.html", context)


def _cabinet_map(projects):
    result = {}
    for project in projects:
        result[str(project.pk)] = [
            {
                "id": cabinet.pk,
                "code": cabinet.cabinet_code,
                "name": cabinet.name,
                "description": cabinet.description,
            }
            for cabinet in project.cabinets.all().order_by("cabinet_code")
        ]
    return result


def _build_cambrics_response(cabinet):
    endpoints = list(
        WireEndpoint.objects.select_related("cabinet", "wire_type", "wire_color")
        .filter(cabinet=cabinet)
        .order_by("wire_color__name", "wire_type__name", "ref", "created_at", "uid")
    )

    columns = defaultdict(lambda: defaultdict(list))
    for endpoint in endpoints:
        if not _endpoint_has_mark(endpoint):
            continue
        columns[_wire_column_label(endpoint)][_ref_group_key(endpoint)].append(endpoint)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Кембрики"

    sheet.append(["Проект", cabinet.project.project_code])
    sheet.append(["Шкаф", cabinet.cabinet_code])
    sheet.append(["Название", cabinet.name])
    sheet.append([])

    column_labels = sorted(columns.keys())
    for column_index, label in enumerate(column_labels, start=1):
        sheet.cell(row=5, column=column_index, value=label)

    group_ranges = []
    max_row = 5
    for column_index, label in enumerate(column_labels, start=1):
        row = 6
        for ref_id in sorted(columns[label].keys()):
            group = sorted(
                columns[label][ref_id],
                key=lambda item: (item.created_at, str(item.uid)),
            )
            mark_labels = _group_mark_labels(group)
            if not mark_labels:
                continue

            start_row = row
            for mark_label in mark_labels:
                sheet.cell(row=row, column=column_index, value=mark_label)
                row += 1
            end_row = row - 1
            if start_row <= end_row:
                group_ranges.append((column_index, start_row, end_row))
            row += 1
        max_row = max(max_row, row - 1)

    _format_cambrics_sheet(sheet, column_labels, group_ranges, max_row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = _cambrics_filename(cabinet)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _wire_column_label(endpoint):
    wire_color = endpoint.wire_color.name if endpoint.wire_color else "-"
    wire_type = endpoint.wire_type.name if endpoint.wire_type else "-"
    return f"{wire_color} - {wire_type}"


def _ref_group_key(endpoint):
    return endpoint.ref or f"NO_REF:{endpoint.uid}"


def _endpoint_has_mark(endpoint):
    return _has_mark_value(endpoint.mark_1) or _has_mark_value(endpoint.mark_2)


def _group_mark_labels(endpoints):
    labels = []
    seen = set()
    for endpoint in endpoints:
        _append_mark_label(labels, seen, endpoint.mark_1)

    if len(labels) < 2:
        for endpoint in endpoints:
            _append_mark_label(labels, seen, endpoint.mark_2)

    return labels


def _append_mark_label(labels, seen, value):
    mark = (value or "").strip()
    if not _has_mark_value(mark) or mark in seen:
        return
    labels.append(mark)
    seen.add(mark)


def _has_mark_value(value):
    mark = (value or "").strip()
    return bool(mark) and mark != "-"


def _format_cambrics_sheet(sheet, column_labels, group_ranges, max_row):
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9D9D9")
    thick_black = Side(style="medium", color="000000")

    for row in range(1, 4):
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=1).fill = title_fill

    header_row = 5
    max_column = max(1, len(column_labels))
    for cell in sheet[header_row][:max_column]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = 24

    for row in sheet.iter_rows(min_row=6, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=thin_gray,
                right=thin_gray,
                top=thin_gray,
                bottom=thin_gray,
            )

    sheet.freeze_panes = "A6"

    for column_index, start_row, end_row in group_ranges:
        for row in range(start_row, end_row + 1):
            cell = sheet.cell(row=row, column=column_index)
            cell.border = Border(
                left=thick_black,
                right=thick_black,
                top=thick_black if row == start_row else thin_gray,
                bottom=thick_black if row == end_row else thin_gray,
            )


def _cambrics_filename(cabinet):
    base = slugify(f"cambrics-{cabinet.project.project_code}-{cabinet.cabinet_code}", allow_unicode=False)
    return f"{base or 'cambrics'}.xlsx"
