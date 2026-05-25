from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.authtoken.models import Token
from rest_framework.test import APIClient

from marking.models import Cabinet, Project, WireEndpoint
from marking.web_views import _build_cambrics_response


class WireEndpointSyncTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="admin", password="awesome1")
        self.token = Token.objects.create(user=self.user)
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.token.key}")
        self.project = Project.objects.create(project_code="001-220", name="Тест")
        self.cabinet = Cabinet.objects.create(
            project=self.project,
            cabinet_code="00CDE02",
            name="Шкаф АСР",
        )

    def test_patch_ref_does_not_clear_marks(self):
        endpoint = self._create_endpoint()

        response = self.client.patch(
            f"/api/wire-endpoints/{endpoint.uid}/",
            {"ref_id": "REF-001", "sync_status": "DIRTY"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        endpoint.refresh_from_db()
        self.assertEqual(endpoint.mark_1, "X1:1")
        self.assertEqual(endpoint.mark_2, "K1:2")
        self.assertEqual(endpoint.ref, "REF-001")

    def test_patch_wire_does_not_clear_marks(self):
        endpoint = self._create_endpoint()

        response = self.client.patch(
            f"/api/wire-endpoints/{endpoint.uid}/",
            {"wire_type": "1,5 кв.мм", "wire_color": "Белый", "sync_status": "DIRTY"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        endpoint.refresh_from_db()
        self.assertEqual(endpoint.mark_1, "X1:1")
        self.assertEqual(endpoint.mark_2, "K1:2")
        self.assertEqual(endpoint.wire_type.name, "1,5 кв.мм")
        self.assertEqual(endpoint.wire_color.name, "Белый")

    def test_patch_status_does_not_clear_marks(self):
        endpoint = self._create_endpoint()

        response = self.client.patch(
            f"/api/wire-endpoints/{endpoint.uid}/",
            {"sync_status": "DIRTY"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        endpoint.refresh_from_db()
        self.assertEqual(endpoint.mark_1, "X1:1")
        self.assertEqual(endpoint.mark_2, "K1:2")
        self.assertEqual(endpoint.sync_status, "DIRTY")

    def test_create_copy_payload_preserves_marks(self):
        response = self.client.post(
            "/api/wire-endpoints/",
            {
                "cabinet": self.cabinet.pk,
                "ref_id": "REF-COPY",
                "mark_1": "X2:10",
                "mark_2": "QF1:2",
                "wire_type": "2,5 кв.мм",
                "wire_color": "Желто-зеленый",
                "sync_status": "SYNCED",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        endpoint = WireEndpoint.objects.get(uid=response.data["endpoint_id"])
        self.assertEqual(endpoint.mark_1, "X2:10")
        self.assertEqual(endpoint.mark_2, "QF1:2")
        self.assertEqual(endpoint.wire_type.name, "2,5 кв.мм")
        self.assertEqual(endpoint.wire_color.name, "Желто-зеленый")

    def test_cambrics_report_uses_marks_after_wire_patch(self):
        endpoint = self._create_endpoint()
        self.client.patch(
            f"/api/wire-endpoints/{endpoint.uid}/",
            {"wire_type": "1,5 кв.мм", "wire_color": "Белый", "sync_status": "DIRTY"},
            format="json",
        )

        response = _build_cambrics_response(self.cabinet)
        sheet = load_workbook(BytesIO(response.content)).active
        values = [
            sheet.cell(row=row, column=1).value
            for row in range(1, sheet.max_row + 1)
        ]

        self.assertIn("Белый - 1,5 кв.мм", values)
        self.assertIn("X1:1", values)
        self.assertIn("K1:2", values)

    def _create_endpoint(self):
        return WireEndpoint.objects.create(
            cabinet=self.cabinet,
            ref="REF-OLD",
            mark_1="X1:1",
            mark_2="K1:2",
            sync_status=WireEndpoint.SyncStatus.SYNCED,
        )
